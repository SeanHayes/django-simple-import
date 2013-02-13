from django.contrib.auth.models import User
from django.contrib.contenttypes.models import ContentType
from django.db import models

class ImportSetting(models.Model):
    """ Save some settings per user per content type """
    user = models.ForeignKey(User)
    content_type = models.ForeignKey(ContentType)
    
class ColumnMatch(models.Model):
    """ Match column names from the user uploaded file to the database """
    column_name = models.CharField(max_length=255)
    field_name = models.CharField(max_length=255)
    import_setting = models.ForeignKey(ImportSetting)
    
class ImportLog(models.Model):
    """ A log of all import attempts """
    name = models.CharField(max_length=255)
    user = models.ForeignKey(User, editable=False, related_name="simple_import_log")
    date = models.DateTimeField(auto_now_add=True)
    import_file = models.FileField(upload_to="import_file")
    error_file = models.FileField(upload_to="error_file", blank=True)
    import_setting = models.ForeignKey(ImportSetting, editable=False)
    def __unicode__(self):
        return unicode(self.name)
    
    def clean(self):
        from django.core.exceptions import ValidationError
        filename = str(self.import_file).lower()
        if not filename[-3:] in ('xls', 'ods', 'csv', 'lsx'):
            raise ValidationError('Invalid file type. Must be xls, xlsx, ods, or csv.')
    
    
    def get_import_file_as_list(self):
        file_ext = str(self.import_file).lower()[-3:]
        data = []
        if file_ext == "xls":
            import xlrd
            wb = xlrd.open_workbook(file_contents=self.import_file.read())
            sh1 = wb.sheet_by_index(0)
            for rownum in range(sh1.nrows): 
                data += [sh1.row_values(rownum)]
        elif file_ext == "csv":
            import csv
            reader = csv.reader(open(self.import_file.path, "rb"))
            for row in reader:
                data += [row]
        elif file_ext == "lsx":
            from openpyxl.reader.excel import load_workbook
            wb = load_workbook(filename=self.import_file.path, use_iterators = True)
            sheet = wb.get_active_sheet()
            for row in sheet.iter_rows():
                data_row = []
                for cell in row:
                    data_row += [cell.internal_value]
                data += [data_row]
        elif file_ext == "ods":
            #TODO add support, all ods libraries suck
            pass